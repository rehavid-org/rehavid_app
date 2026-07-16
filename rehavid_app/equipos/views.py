"""Vistas del inventario de equipos (Fase 4). Mutaciones vía
``reservas/services.py`` (listo/mantenimiento/baja) y auditoría."""

from collections import Counter

from django.contrib import messages
from django.shortcuts import get_object_or_404
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.views.generic import CreateView
from django.views.generic import ListView
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from rehavid_app.auditoria import services as auditoria
from rehavid_app.catalogo.models import Ciudad
from rehavid_app.catalogo.models import Servicio
from rehavid_app.reservas import services as reservas_service
from rehavid_app.reservas.services import ReservaError
from rehavid_app.users.permissions import NivelRequeridoMixin
from rehavid_app.users.permissions import nivel_requerido
from rehavid_app.xlsx import workbook_response

from .forms import BajaForm
from .forms import EquipoForm
from .forms import ListoForm
from .forms import MantenimientoForm
from .models import Equipo
from .models import EstadoEquipo


class EquipoListView(NivelRequeridoMixin, ListView):
    nivel_maximo = 3
    model = Equipo
    template_name = "equipos/lista.html"
    context_object_name = "equipos"

    def get_queryset(self):
        qs = Equipo.objects.select_related("servicio", "ciudad_base").prefetch_related("accesorios")
        p = self.request.GET
        if q := p.get("q", "").strip():
            qs = qs.filter(codigo__icontains=q) | qs.filter(serial__icontains=q) | qs.filter(modelo__icontains=q)
        if servicio := p.get("servicio"):
            qs = qs.filter(servicio_id=servicio)
        if estado := p.get("estado"):
            qs = qs.filter(estado=estado)
        return qs.distinct().order_by("codigo")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        inventario = Equipo.objects.exclude(estado=EstadoEquipo.DE_BAJA)
        ctx.update(
            modulo_activo="equipos",
            # O02 · KPI disponibles/total en vivo
            kpi_total=inventario.count(),
            kpi_disponibles=inventario.filter(estado=EstadoEquipo.DISPONIBLE).count(),
            kpi_en_uso=inventario.filter(estado=EstadoEquipo.EN_USO).count(),
            kpi_preparacion=inventario.filter(estado=EstadoEquipo.EN_PREPARACION).count(),
            # O07 · Tumeke y demás servicios sin unidad física, como tarjeta especial
            servicios_sin_stock=Servicio.objects.filter(requiere_equipo_fisico=False, activo=True),
            servicios=Servicio.objects.filter(activo=True, requiere_equipo_fisico=True),
            estados=EstadoEquipo.choices,
            filtros=self.request.GET,
            listo_form=ListoForm(),
            mantenimiento_form=MantenimientoForm(),
            baja_form=BajaForm(),
            stock_por_categoria=self._stock_por_categoria(inventario),
            equipos_por_ciudad=self._equipos_por_ciudad(inventario),
        )
        return ctx

    @staticmethod
    def _stock_por_categoria(inventario):
        por_categoria: dict[str, dict[str, int]] = {}
        for e in inventario.select_related("servicio"):
            fila = por_categoria.setdefault(e.servicio.nombre, {"total": 0, "disponibles": 0})
            fila["total"] += 1
            if e.estado == EstadoEquipo.DISPONIBLE:
                fila["disponibles"] += 1
        categorias = sorted(por_categoria)
        return {
            "categorias": categorias,
            "total": [por_categoria[c]["total"] for c in categorias],
            "disponibles": [por_categoria[c]["disponibles"] for c in categorias],
        }

    @staticmethod
    def _equipos_por_ciudad(inventario):
        conteo = Counter(inventario.values_list("ciudad_base__nombre", flat=True))
        ciudades = sorted(conteo, key=lambda c: conteo[c])
        return {"ciudades": ciudades, "conteos": [conteo[c] for c in ciudades]}


class EquipoCreateView(NivelRequeridoMixin, CreateView):
    """B7 · alta con el modelo canónico; aparece en inventario y disponibilidad."""

    nivel_maximo = 2
    model = Equipo
    form_class = EquipoForm
    template_name = "equipos/alta.html"
    success_url = reverse_lazy("equipos:lista")

    def form_valid(self, form):
        response = super().form_valid(form)
        auditoria.registrar(
            self.request.user, "crear_equipo", "equipos",
            f"{self.object.codigo} · {self.object.modelo} · serial {self.object.serial}",
        )
        messages.success(self.request, f"Equipo {self.object.codigo} agregado al inventario")
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["modulo_activo"] = "equipos"
        return ctx


COLUMNAS_IMPORT = ["codigo", "servicio", "modelo", "serial", "ciudad_base", "responsable", "notas"]


@nivel_requerido(2)
def export_view(request):
    filas = [
        [
            e.codigo,
            e.servicio.nombre,
            e.modelo,
            e.serial,
            e.ciudad_base.nombre,
            e.get_estado_display(),
            e.responsable,
            e.historial_uso,
            e.ultima_revision,
            e.proxima_mantencion,
            e.notas,
        ]
        for e in Equipo.objects.select_related("servicio", "ciudad_base").order_by("codigo")
    ]
    auditoria.registrar(request.user, "export_equipos", "equipos", f"{len(filas)} filas")
    return workbook_response(
        "equipos_rehavid.xlsx",
        "Equipos",
        ["Código", "Categoría", "Modelo", "Serial", "Ciudad base", "Estado",
         "Responsable", "Usos", "Última revisión", "Próx. mantención", "Notas"],
        filas,
    )


@nivel_requerido(2)
def plantilla_import_view(request):
    """Plantilla con las columnas EXACTAS del modelo canónico (B7)."""
    ejemplo = [["XS-99", "Xsens", "Xsens MVN Link", "SN-XS-099", "Medellín", "Bodega Medellín", ""]]
    return workbook_response("plantilla_equipos.xlsx", "Plantilla equipos", COLUMNAS_IMPORT, ejemplo)


def _validar_fila_import(fila, servicios: dict, ciudades: dict) -> tuple[Equipo | None, str | None]:
    """Valida una fila de la plantilla contra el modelo canónico (B7)."""
    codigo, servicio_n, modelo, serial, ciudad_n, responsable, notas = (
        str(c).strip() if c is not None else "" for c in (list(fila) + [""] * 7)[:7]
    )
    if not all([codigo, servicio_n, modelo, serial, ciudad_n]):
        return None, "codigo/servicio/modelo/serial/ciudad_base son obligatorios"
    servicio = servicios.get(servicio_n.lower())
    if servicio is None:
        return None, f"servicio '{servicio_n}' no existe en el catálogo"
    ciudad = ciudades.get(ciudad_n.lower())
    if ciudad is None:
        return None, f"ciudad '{ciudad_n}' no existe en el catálogo"
    if Equipo.objects.filter(codigo=codigo).exists() or Equipo.objects.filter(serial=serial).exists():
        return None, f"código {codigo} o serial {serial} ya existen en el inventario"
    equipo = Equipo(
        codigo=codigo, servicio=servicio, modelo=modelo, serial=serial,
        ciudad_base=ciudad, responsable=responsable, notas=notas,
    )
    return equipo, None


@nivel_requerido(2)
def import_view(request):
    """B7/B14 · import validado contra el modelo canónico; todo-o-nada."""
    archivo = request.FILES.get("archivo")
    if not archivo:
        messages.error(request, "Adjunte el archivo .xlsx (use la plantilla)")
        return redirect("equipos:lista")

    try:
        wb = load_workbook(archivo, read_only=True, data_only=True)
    except (InvalidFileException, KeyError, OSError):
        messages.error(request, "El archivo no es un .xlsx válido")
        return redirect("equipos:lista")
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas or [str(c or "").strip().lower() for c in filas[0][: len(COLUMNAS_IMPORT)]] != COLUMNAS_IMPORT:
        messages.error(request, f"Encabezados inválidos. Use la plantilla: {', '.join(COLUMNAS_IMPORT)}")
        return redirect("equipos:lista")

    servicios = {s.nombre.lower(): s for s in Servicio.objects.all()}
    ciudades = {c.nombre.lower(): c for c in Ciudad.objects.all()}
    errores, nuevos = [], []
    for idx, fila in enumerate(filas[1:], start=2):
        if not any(fila):
            continue
        equipo, error = _validar_fila_import(fila, servicios, ciudades)
        if error:
            errores.append(f"Fila {idx}: {error}")
        else:
            nuevos.append(equipo)

    if errores:
        messages.error(request, "Import rechazado (nada se creó): " + " · ".join(errores[:8]))
        return redirect("equipos:lista")
    Equipo.objects.bulk_create(nuevos)
    auditoria.registrar(request.user, "import_equipos", "equipos", f"{len(nuevos)} equipos importados")
    messages.success(request, f"{len(nuevos)} equipo(s) importados al inventario")
    return redirect("equipos:lista")


@nivel_requerido(2)
def listo_view(request, pk):
    equipo = get_object_or_404(Equipo, pk=pk)
    form = ListoForm(request.POST)
    notas = form.cleaned_data["notas"] if form.is_valid() else ""
    try:
        reservas_service.marcar_equipo_listo(equipo, notas, request.user)
        messages.success(request, f"{equipo.codigo} listo y disponible")
    except ReservaError as e:
        messages.error(request, str(e))
    return redirect("equipos:lista")


@nivel_requerido(2)
def mantenimiento_view(request, pk):
    equipo = get_object_or_404(Equipo, pk=pk)
    form = MantenimientoForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Indique el motivo del mantenimiento")
        return redirect("equipos:lista")
    reservas_service.enviar_a_mantenimiento(equipo, form.cleaned_data["motivo"], request.user)
    messages.success(request, f"{equipo.codigo} enviado a mantenimiento")
    return redirect("equipos:lista")


@nivel_requerido(1)  # O18 · baja definitiva solo Admin Global
def baja_view(request, pk):
    equipo = get_object_or_404(Equipo, pk=pk)
    form = BajaForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Indique el motivo de la baja")
        return redirect("equipos:lista")
    try:
        reservas_service.dar_de_baja_equipo(equipo, form.cleaned_data["motivo"], request.user)
        messages.success(request, f"{equipo.codigo} dado de baja definitivamente")
    except ReservaError as e:
        messages.error(request, str(e))
    return redirect("equipos:lista")
