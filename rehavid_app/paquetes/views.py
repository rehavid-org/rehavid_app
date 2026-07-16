"""Paquetes multi-equipo: tarjetas tri-estado (O09) + CRUD nivel <= 2 (O20/B11)."""

from datetime import timedelta

from django.contrib import messages
from django.http import HttpResponseRedirect
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import CreateView
from django.views.generic import DeleteView
from django.views.generic import ListView
from django.views.generic import UpdateView
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from rehavid_app.auditoria import services as auditoria
from rehavid_app.catalogo.models import Servicio
from rehavid_app.reservas import services as reservas_service
from rehavid_app.users.permissions import NivelRequeridoMixin
from rehavid_app.users.permissions import nivel_requerido
from rehavid_app.xlsx import workbook_response

from .forms import PaqueteForm
from .models import Paquete

COLUMNAS_IMPORT = ["codigo", "nombre", "descripcion", "servicios_requeridos", "duracion_dias"]


def estado_paquete(paquete: Paquete) -> dict:
    """O09 · tri-estado hoy→hoy+duración: disponible / parcial / agotado."""
    hoy = timezone.localdate()
    fin = hoy + timedelta(days=max(paquete.duracion_dias - 1, 0))
    v = reservas_service.verificar_disponibilidad_paquete(paquete, hoy, fin)
    libres = sum(1 for d in v["detalle"] if d["disponible"])
    total = len(v["detalle"])
    if v["disponible"]:
        estado = "disponible"
    elif libres:
        estado = "parcial"
    else:
        estado = "agotado"

    proxima = None
    if estado != "disponible":
        # Estimación: la fecha en que la última categoría faltante vuelve a tener stock
        fechas = [
            reservas_service.proxima_fecha_disponible(servicio, paquete.duracion_dias)
            for servicio, d in zip(paquete.servicios_requeridos.all(), v["detalle"], strict=False)
            if not d["disponible"]
        ]
        fechas = [f for f in fechas if f]
        proxima = max(fechas) if fechas else None

    # O09 · usos reales en los últimos 30 días. En el prototipo origen este número
    # era un campo fijo del mock, nunca recalculado al crear nuevas reservas.
    hace_30d = hoy - timedelta(days=30)
    uso_30d = paquete.reservas.filter(cancelada=False, fecha_salida__gte=hace_30d).count()

    return {
        "paquete": paquete,
        "estado": estado,
        "motivo": v["motivo"],
        "detalle": v["detalle"],
        "libres": libres,
        "total": total,
        "proxima": proxima,
        "uso_30d": uso_30d,
    }


class PaqueteListView(NivelRequeridoMixin, ListView):
    nivel_maximo = 3
    model = Paquete
    template_name = "paquetes/lista.html"
    context_object_name = "paquetes"

    def get_queryset(self):
        tab = self.request.GET.get("tab", "activos")
        return Paquete.objects.filter(activo=(tab != "inactivos")).prefetch_related("servicios_requeridos")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        tarjetas = [estado_paquete(p) for p in ctx["paquetes"]]
        if self.request.GET.get("tab") == "mas_usados":
            tarjetas.sort(key=lambda t: t["uso_30d"], reverse=True)
        ctx.update(
            modulo_activo="paquetes",
            tarjetas=tarjetas,
            tab=self.request.GET.get("tab", "activos"),
            kpi_activos=Paquete.objects.filter(activo=True).count(),
            kpi_inactivos=Paquete.objects.filter(activo=False).count(),
        )
        return ctx


class PaqueteCreateView(NivelRequeridoMixin, CreateView):
    nivel_maximo = 2
    model = Paquete
    form_class = PaqueteForm
    template_name = "paquetes/form.html"
    success_url = reverse_lazy("paquetes:lista")

    def form_valid(self, form):
        response = super().form_valid(form)
        auditoria.registrar(self.request.user, "crear_paquete", "paquetes", self.object.codigo)
        messages.success(self.request, f"Paquete {self.object.codigo} creado")
        return response

    def get_context_data(self, **kwargs):
        return super().get_context_data(**kwargs) | {"modulo_activo": "paquetes", "titulo": "Nuevo paquete"}


class PaqueteUpdateView(NivelRequeridoMixin, UpdateView):
    nivel_maximo = 2
    model = Paquete
    form_class = PaqueteForm
    template_name = "paquetes/form.html"
    success_url = reverse_lazy("paquetes:lista")

    def form_valid(self, form):
        response = super().form_valid(form)
        auditoria.registrar(self.request.user, "editar_paquete", "paquetes", self.object.codigo)
        messages.success(self.request, f"Paquete {self.object.codigo} actualizado")
        return response

    def get_context_data(self, **kwargs):
        return super().get_context_data(**kwargs) | {
            "modulo_activo": "paquetes",
            "titulo": f"Editar {self.object.codigo}",
        }


class PaqueteDeleteView(NivelRequeridoMixin, DeleteView):
    nivel_maximo = 2
    model = Paquete
    template_name = "paquetes/eliminar.html"
    success_url = reverse_lazy("paquetes:lista")

    def form_valid(self, form):
        codigo = self.object.codigo
        if self.object.reservas.exists():
            # Con historial de reservas no se borra: se desactiva
            self.object.activo = False
            self.object.save(update_fields=["activo"])
            auditoria.registrar(self.request.user, "desactivar_paquete", "paquetes", codigo)
            messages.success(self.request, f"Paquete {codigo} desactivado (tiene reservas históricas)")
            return HttpResponseRedirect(self.get_success_url())
        auditoria.registrar(self.request.user, "eliminar_paquete", "paquetes", codigo)
        messages.success(self.request, f"Paquete {codigo} eliminado")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        return super().get_context_data(**kwargs) | {"modulo_activo": "paquetes"}


@nivel_requerido(2)
def export_view(request):
    filas = [
        [
            p.codigo,
            p.nombre,
            p.descripcion,
            ", ".join(s.nombre for s in p.servicios_requeridos.all()),
            p.duracion_dias,
            "Sí" if p.activo else "No",
            estado_paquete(p)["uso_30d"],
        ]
        for p in Paquete.objects.prefetch_related("servicios_requeridos").order_by("codigo")
    ]
    auditoria.registrar(request.user, "export_paquetes", "paquetes", f"{len(filas)} filas")
    return workbook_response(
        "paquetes_rehavid.xlsx",
        "Paquetes",
        ["Código", "Nombre", "Descripción", "Servicios requeridos", "Duración (días)", "Activo", "Usos 30 días"],
        filas,
    )


@nivel_requerido(2)
def plantilla_import_view(request):
    """Plantilla con las columnas del modelo canónico de Paquete (B7-style)."""
    ejemplo = [["PKG-99", "Pack de ejemplo", "Descripción del paquete", "Xsens, EMG", 2]]
    return workbook_response("plantilla_paquetes.xlsx", "Plantilla paquetes", COLUMNAS_IMPORT, ejemplo)


def _validar_fila_paquete(fila, servicios: dict) -> tuple[Paquete | None, list | None, str | None]:
    """Valida una fila de la plantilla contra el modelo canónico de Paquete."""
    codigo, nombre, descripcion, servicios_raw, duracion_raw = (
        str(c).strip() if c is not None else "" for c in (list(fila) + [""] * 5)[:5]
    )
    if not all([codigo, nombre, servicios_raw]):
        return None, None, "codigo/nombre/servicios_requeridos son obligatorios"
    if Paquete.objects.filter(codigo=codigo).exists():
        return None, None, f"código {codigo} ya existe"
    nombres_servicio = [s.strip() for s in servicios_raw.split(",") if s.strip()]
    servicios_obj = []
    for nombre_servicio in nombres_servicio:
        servicio = servicios.get(nombre_servicio.lower())
        if servicio is None:
            return None, None, f"servicio '{nombre_servicio}' no existe en el catálogo"
        servicios_obj.append(servicio)
    try:
        duracion = int(duracion_raw or 1)
    except ValueError:
        return None, None, "duracion_dias debe ser un número entero"
    paquete = Paquete(codigo=codigo, nombre=nombre, descripcion=descripcion, duracion_dias=duracion)
    return paquete, servicios_obj, None


@nivel_requerido(2)
def import_view(request):
    """Import validado todo-o-nada, mismo patrón que equipos (B7/B14)."""
    archivo = request.FILES.get("archivo")
    if not archivo:
        messages.error(request, "Adjunte el archivo .xlsx (use la plantilla)")
        return redirect("paquetes:lista")

    try:
        wb = load_workbook(archivo, read_only=True, data_only=True)
    except (InvalidFileException, KeyError, OSError):
        messages.error(request, "El archivo no es un .xlsx válido")
        return redirect("paquetes:lista")
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas or [str(c or "").strip().lower() for c in filas[0][: len(COLUMNAS_IMPORT)]] != COLUMNAS_IMPORT:
        messages.error(request, f"Encabezados inválidos. Use la plantilla: {', '.join(COLUMNAS_IMPORT)}")
        return redirect("paquetes:lista")

    servicios = {s.nombre.lower(): s for s in Servicio.objects.all()}
    errores, nuevos = [], []
    for idx, fila in enumerate(filas[1:], start=2):
        if not any(fila):
            continue
        paquete, servicios_obj, error = _validar_fila_paquete(fila, servicios)
        if error:
            errores.append(f"Fila {idx}: {error}")
        else:
            nuevos.append((paquete, servicios_obj))

    if errores:
        messages.error(request, "Import rechazado (nada se creó): " + " · ".join(errores[:8]))
        return redirect("paquetes:lista")
    for paquete, servicios_obj in nuevos:
        paquete.save()
        paquete.servicios_requeridos.set(servicios_obj)
    auditoria.registrar(request.user, "import_paquetes", "paquetes", f"{len(nuevos)} paquetes importados")
    messages.success(request, f"{len(nuevos)} paquete(s) importados")
    return redirect("paquetes:lista")
