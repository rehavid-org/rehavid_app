"""CRUD de planes de acción (B11). Nivel <= 2; lectura para quien tenga el módulo."""

from collections import Counter
from datetime import date
from datetime import datetime

from django.contrib import messages
from django.db.models import Avg
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import CreateView
from django.views.generic import DeleteView
from django.views.generic import ListView
from django.views.generic import UpdateView
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from rehavid_app.auditoria import services as auditoria
from rehavid_app.users.permissions import ModuloRequeridoMixin
from rehavid_app.users.permissions import NivelRequeridoMixin
from rehavid_app.users.permissions import nivel_requerido
from rehavid_app.xlsx import workbook_response

from .forms import PlanForm
from .models import Plan

COLUMNAS_IMPORT = ["area", "titulo", "descripcion", "responsable", "vence", "avance", "esperado", "estado"]
ESTADOS_VALIDOS = {e for e, _ in Plan.Estado.choices}


class PlanListView(ModuloRequeridoMixin, ListView):
    modulo = "planes"
    model = Plan
    template_name = "planes/lista.html"
    context_object_name = "planes"

    def get_queryset(self):
        qs = Plan.objects.all()
        if estado := self.request.GET.get("estado"):
            qs = qs.filter(estado=estado)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(
            modulo_activo="planes",
            filtros=self.request.GET,
            estados=Plan.Estado.choices,
            kpi_abiertos=Plan.objects.filter(estado=Plan.Estado.ABIERTO).count(),
            kpi_riesgo=Plan.objects.filter(estado=Plan.Estado.EN_RIESGO).count(),
            kpi_completados=Plan.objects.filter(estado=Plan.Estado.COMPLETADO).count(),
            resumen=self._resumen(),
        )
        return ctx

    @staticmethod
    def _resumen():
        """Panel lateral: en el prototipo origen estos totales, el avance
        promedio y el desglose "apps origen" eran texto fijo, nunca
        recalculado ni al importar planes por Excel; acá salen de la BD."""
        planes = list(Plan.objects.all())
        total = len(planes)
        en_riesgo = [p for p in planes if p.estado == Plan.Estado.EN_RIESGO]
        avance_promedio = round(Plan.objects.aggregate(a=Avg("avance"))["a"] or 0)
        hoy = timezone.localdate()
        proximos = [p for p in planes if p.estado != Plan.Estado.COMPLETADO and p.vence >= hoy]
        proximo_vencimiento = min((p.vence - hoy).days for p in proximos) if proximos else None
        por_area = Counter(p.area for p in planes)
        return {
            "total": total,
            "avance_promedio": avance_promedio,
            "proximo_vencimiento_dias": proximo_vencimiento,
            "por_area": sorted(por_area.items(), key=lambda kv: kv[1], reverse=True),
            "en_riesgo": en_riesgo,
        }


class PlanCreateView(NivelRequeridoMixin, CreateView):
    nivel_maximo = 2
    model = Plan
    form_class = PlanForm
    template_name = "planes/form.html"
    success_url = reverse_lazy("planes:lista")

    def form_valid(self, form):
        response = super().form_valid(form)
        auditoria.registrar(self.request.user, "crear_plan", "planes", self.object.codigo)
        messages.success(self.request, f"Plan {self.object.codigo} creado")
        return response

    def get_context_data(self, **kwargs):
        return super().get_context_data(**kwargs) | {"modulo_activo": "planes", "titulo": "Nuevo plan de acción"}


class PlanUpdateView(NivelRequeridoMixin, UpdateView):
    nivel_maximo = 2
    model = Plan
    form_class = PlanForm
    template_name = "planes/form.html"
    success_url = reverse_lazy("planes:lista")

    def form_valid(self, form):
        response = super().form_valid(form)
        auditoria.registrar(self.request.user, "editar_plan", "planes", self.object.codigo)
        messages.success(self.request, f"Plan {self.object.codigo} actualizado")
        return response

    def get_context_data(self, **kwargs):
        return super().get_context_data(**kwargs) | {
            "modulo_activo": "planes",
            "titulo": f"Editar {self.object.codigo}",
        }


class PlanDeleteView(NivelRequeridoMixin, DeleteView):
    nivel_maximo = 2
    model = Plan
    template_name = "planes/eliminar.html"
    success_url = reverse_lazy("planes:lista")

    def form_valid(self, form):
        codigo = self.object.codigo
        auditoria.registrar(self.request.user, "eliminar_plan", "planes", codigo)
        messages.success(self.request, f"Plan {codigo} eliminado")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        return super().get_context_data(**kwargs) | {"modulo_activo": "planes"}


@nivel_requerido(2)
def export_view(request):
    filas = [
        [
            p.codigo, p.area, p.titulo, p.responsable, p.vence,
            p.avance, p.esperado, p.get_estado_display(),
        ]
        for p in Plan.objects.all()
    ]
    auditoria.registrar(request.user, "export_planes", "planes", f"{len(filas)} filas")
    return workbook_response(
        "planes_rehavid.xlsx",
        "Planes",
        ["Código", "Área", "Título", "Responsable", "Vence", "Avance %", "Esperado %", "Estado"],
        filas,
    )


@nivel_requerido(2)
def plantilla_import_view(request):
    """Plantilla con las columnas del modelo canónico de Plan (B7-style)."""
    ejemplo = [[
        "Operaciones · Reservas", "Protocolo de retorno temprano", "Activar contacto 48h antes del retorno",
        "Ariel Ramírez", "2026-08-15", 0, 20, "open",
    ]]
    return workbook_response("plantilla_planes.xlsx", "Plantilla planes", COLUMNAS_IMPORT, ejemplo)


def _parsear_fecha(valor) -> date | None:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    try:
        return datetime.strptime(str(valor).strip(), "%Y-%m-%d").date()  # noqa: DTZ007
    except ValueError:
        return None


def _validar_fila_plan(fila) -> tuple[Plan | None, str | None]:
    """Valida una fila de la plantilla contra el modelo canónico de Plan."""
    area, titulo, descripcion, responsable, vence_raw, avance_raw, esperado_raw, estado = (
        str(c).strip() if c is not None else "" for c in (list(fila) + [""] * 8)[:8]
    )
    if not all([area, titulo, responsable, vence_raw, estado]):
        return None, "area/titulo/responsable/vence/estado son obligatorios"
    vence = _parsear_fecha(fila[4])
    if vence is None:
        return None, f"fecha '{vence_raw}' inválida (use YYYY-MM-DD)"
    try:
        avance = int(avance_raw or 0)
        esperado = int(esperado_raw or 0)
    except ValueError:
        return None, "avance/esperado deben ser números enteros"
    if not (0 <= avance <= 100) or not (0 <= esperado <= 100):  # noqa: PLR2004
        return None, "avance/esperado deben estar entre 0 y 100"
    if estado.lower() not in ESTADOS_VALIDOS:
        return None, f"estado '{estado}' inválido (use: {', '.join(ESTADOS_VALIDOS)})"
    plan = Plan(
        area=area, titulo=titulo, descripcion=descripcion, responsable=responsable,
        vence=vence, avance=avance, esperado=esperado, estado=estado.lower(),
    )
    return plan, None


@nivel_requerido(2)
def import_view(request):
    """Import validado todo-o-nada, mismo patrón que equipos (B7/B14)."""
    archivo = request.FILES.get("archivo")
    if not archivo:
        messages.error(request, "Adjunte el archivo .xlsx (use la plantilla)")
        return redirect("planes:lista")

    try:
        wb = load_workbook(archivo, read_only=True, data_only=True)
    except (InvalidFileException, KeyError, OSError):
        messages.error(request, "El archivo no es un .xlsx válido")
        return redirect("planes:lista")
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas or [str(c or "").strip().lower() for c in filas[0][: len(COLUMNAS_IMPORT)]] != COLUMNAS_IMPORT:
        messages.error(request, f"Encabezados inválidos. Use la plantilla: {', '.join(COLUMNAS_IMPORT)}")
        return redirect("planes:lista")

    errores, nuevos = [], []
    for idx, fila in enumerate(filas[1:], start=2):
        if not any(fila):
            continue
        plan, error = _validar_fila_plan(fila)
        if error:
            errores.append(f"Fila {idx}: {error}")
        else:
            nuevos.append(plan)

    if errores:
        messages.error(request, "Import rechazado (nada se creó): " + " · ".join(errores[:8]))
        return redirect("planes:lista")
    for plan in nuevos:
        plan.save()  # uno a uno: cada save() asigna su código PL-### (no admite bulk_create)
    auditoria.registrar(request.user, "import_planes", "planes", f"{len(nuevos)} planes importados")
    messages.success(request, f"{len(nuevos)} plan(es) importados")
    return redirect("planes:lista")
