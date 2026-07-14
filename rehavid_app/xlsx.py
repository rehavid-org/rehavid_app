"""Helper openpyxl compartido: reemplaza el export SheetJS del prototipo
por export server-side auditable (B14)."""

from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
ENCABEZADO_FILL = PatternFill("solid", fgColor="4025CE")  # morado Rehavid
ENCABEZADO_FONT = Font(color="FFFFFF", bold=True)


def workbook_response(nombre_archivo: str, titulo: str, encabezados: list[str], filas: list[list]) -> HttpResponse:
    """Arma un .xlsx de una hoja con encabezado con estilo de marca."""
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    ws.append(encabezados)
    for celda in ws[1]:
        celda.fill = ENCABEZADO_FILL
        celda.font = ENCABEZADO_FONT
    for fila in filas:
        ws.append(fila)

    for idx, encabezado in enumerate(encabezados, start=1):
        largos = [len(str(f[idx - 1])) for f in filas if f[idx - 1] is not None][:200]
        largo = max([len(str(encabezado)), *largos])
        ws.column_dimensions[get_column_letter(idx)].width = min(largo + 3, 45)
    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(buffer.getvalue(), content_type=MIME_XLSX)
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response
